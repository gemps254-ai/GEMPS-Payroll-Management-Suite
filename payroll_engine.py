import streamlit as st
import pandas as pd
import os
from datetime import datetime
import io
import zipfile
from fpdf import FPDF

st.set_page_config(page_title="Fahiri 🇰🇪 Payroll Pro", layout="wide")

# --- 1. THE PURE PYTHON CALCULATION ENGINE (REPLACES EXCEL) ---
def run_payroll_calculations(df):
    results = []
    for _, row in df.iterrows():
        # Inputs from the Queue
        basic = float(row["Basic Salary"])
        extra = float(row["Other Allowances"])
        pension_contrib = float(row["Pension Contribution"])
        
        # Gross Pay
        gross_pay = basic + extra
        
        # NSSF Calculation
        # Using Feb 2026 Tier II cap (108,000 * 6%)
        nssf_deduction = min(gross_pay * 0.06, 108000 * 0.06)
        
        # SHIF & AHL
        shif = gross_pay * 0.0275
        housing_levy = gross_pay * 0.015
        
        # Taxable Salary
        # Standard Kenyan logic: Gross - NSSF - SHIF - Housing Levy - Allowable Pension (capped at 30k)
        p_limit = 30000
        allowable_pension = min(pension_contrib + nssf_deduction, p_limit)
        taxable_salary = gross_pay - nssf_deduction - shif - housing_levy - allowable_pension
        
        # PAYE Brackets
        tax = 0
        rem = taxable_salary
        # KRA Tax Brackets
        if rem > 0:
            # 10% on first 24,000
            b1 = min(rem, 24000)
            tax += b1 * 0.10
            rem -= b1
        if rem > 0:
            # 25% on next 8,333
            b2 = min(rem, 32333 - 24000)
            tax += b2 * 0.25
            rem -= b2
        if rem > 0:
            # 30% on next 467,667
            b3 = min(rem, 500000 - 32333)
            tax += b3 * 0.30
            rem -= b3
        if rem > 0:
            # 32.5% on next 300,000
            b4 = min(rem, 800000 - 500000)
            tax += b4 * 0.325
            rem -= b4
        if rem > 0:
            # 35% on anything above 800k
            tax += rem * 0.35
            
        # Personal Relief
        personal_relief = 2400.00
        paye = max(0, tax - personal_relief)
        
        # Net Pay
        net_pay = gross_pay - pension_contrib - nssf_deduction - shif - housing_levy - paye
        
        results.append({
            "Staff No": row["Staff_No."],
            "Employee": row["Employee Name"],
            "Basic Salary": basic,
            "Allowances": extra,
            "Pension": pension_contrib,
            "Gross Pay (KES)": gross_pay,
            "NSSF": nssf_deduction,
            "SHIF": shif,
            "Housing Levy": housing_levy,
            "Taxable Salary": taxable_salary,
            "Personal Relief": personal_relief,
            "PAYE": paye,
            "Net Pay (KES)": net_pay
        })
    return pd.DataFrame(results)

# --- PDF PAYSLIP GENERATOR (RETAINED) ---
class Payslip(FPDF):
    def header(self):
        self.set_font('Arial', 'B', 17)
        self.ln(15)
        self.cell(0, 15, 'FAHIRI KE LIMITED', 0, 1, 'C')
        self.set_font('Arial', '', 12)
        self.cell(0, 5, 'Official Employee Payslip', 0, 1, 'C')
        self.ln(5)

def generate_payslip_pdf(emp_data, month, year):
    pdf = Payslip()
    pdf.add_page()
    pdf.set_font("Arial", size=11)
    pdf.cell(200, 10, txt=f"Payroll Period: {month} {year}", ln=True, align='L')
    pdf.cell(200, 10, txt=f"Staff Number: {emp_data['Staff No']}", ln=True, align='L')
    pdf.cell(200, 10, txt=f"Employee Name: {emp_data['Employee']}", ln=True, align='L')
    pdf.ln(10)
    pdf.set_fill_color(200, 220, 255)
    pdf.cell(95, 10, "Description", 1, 0, 'C', True)
    pdf.cell(95, 10, "Amount (KES)", 1, 1, 'C', True)
    pdf.set_fill_color(240, 240, 240)
    
    data_points = [
        ("Basic Salary", emp_data['Basic Salary']),
        ("Allowances", emp_data['Allowances']),
        ("Gross Pay", emp_data['Gross Pay (KES)']),
        ("Pension Contribution", emp_data['Pension']),
        ("NSSF", emp_data['NSSF']),
        ("SHIF", emp_data['SHIF']),
        ("Housing Levy", emp_data['Housing Levy']),
        ("Taxable Salary", emp_data['Taxable Salary']),
        ("Personal Relief", emp_data['Personal Relief']),
        ("PAYE Deduction", -emp_data['PAYE']),
    ]
    for desc, val in data_points:
        pdf.cell(95, 8, desc, 1)
        pdf.cell(95, 8, f"{val:,.2f}", 1, 1, 'R')
    pdf.set_font("Arial", 'B', 12)
    pdf.cell(95, 10, "NET PAY", 1)
    pdf.cell(95, 10, f"{emp_data['Net Pay (KES)']:,.2f}", 1, 1, 'R')
    return pdf.output(dest='S')

# --- APP LAYOUT (RETAINED) ---
st.title("Fahiri 🇰🇪 Payroll Management Suite")

with st.sidebar:
    st.header("Payroll Data")
    payroll_month = st.selectbox("Month", ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"], index=datetime.now().month - 1)
    payroll_year = st.text_input("Year", value="2026")
    uploaded_file = st.file_uploader("Upload Staff List", type=["csv", "xlsx"])

if "employees" not in st.session_state:
    st.session_state["employees"] = pd.DataFrame(columns=["Staff_No.", "Employee Name", "Basic Salary", "Other Allowances", "Pension Contribution"])
if "results_df" not in st.session_state:
    st.session_state["results_df"] = None

if uploaded_file:
    if st.sidebar.button("Load Data"):
        df_upload = pd.read_csv(uploaded_file) if uploaded_file.name.endswith('.csv') else pd.read_excel(uploaded_file)
        column_mapping = {"staff_no": "Staff_No.", "name": "Employee Name", "basic": "Basic Salary", "extra": "Other Allowances", "pension": "Pension Contribution"}
        st.session_state["employees"] = df_upload.rename(columns=column_mapping)
        st.toast(f"✅ Successfully uploaded {len(df_upload)} employees!", icon='🚀')

with st.expander("Add Employee Manually"):
    with st.form("employee_form", clear_on_submit=True):
        col1, col2 = st.columns(2)
        staff_no = col1.text_input("Staff Number")
        name = col2.text_input("Employee Name")
        basic = col1.number_input("Basic Salary", min_value=0.0, step=100.0)
        extra = col2.number_input("Other Allowances", min_value=0.0, step=100.0)
        pension = col1.number_input("Pension Contribution", min_value=0.0, step=100.0)
        if st.form_submit_button("Add to List"):
            new_row = pd.DataFrame([{"Staff_No.": staff_no, "Employee Name": name, "Basic Salary": basic, "Other Allowances": extra, "Pension Contribution": pension}])
            st.session_state["employees"] = pd.concat([st.session_state["employees"], new_row], ignore_index=True)

st.subheader("📋 Payroll Queue")
st.session_state["employees"] = st.data_editor(st.session_state["employees"], num_rows="dynamic", use_container_width=True)

# --- UPDATED CALCULATION ENGINE TRIGGER ---
if st.button("🚀 Run Full Payroll & Generate Payslips", type="primary"):
    if st.session_state["employees"].empty:
        st.error("Queue is empty!")
    else:
        with st.status("Calculating Payroll...", expanded=True) as status:
            # Perform calculations in pure Python
            st.session_state["results_df"] = run_payroll_calculations(st.session_state["employees"])
            status.update(label="Calculations Complete!", state="complete")

# --- RESULTS DISPLAY (RETAINED) ---
if st.session_state["results_df"] is not None:
    res_df = st.session_state["results_df"]
    st.divider()
    c1, c2, c3 = st.columns(3)
    c1.metric("Bank Transfer for Salaries", f"KES {res_df['Net Pay (KES)'].sum():,.2f}")
    c2.metric("PAYE to be Remitted (KRA)", f"KES {res_df['PAYE'].sum():,.2f}")
    c3.metric("Salaries Processed", len(res_df))

    import io
import openpyxl
from openpyxl.utils import get_column_letter

st.subheader("📦 Export Files")
d1, d2 = st.columns(2)

with d1:
    # 1. Load your existing template from the repository
    template_path = "Payroll Summary Template.xlsx"
    
    try:
        wb = openpyxl.load_workbook(template_path)
        sheet = wb["Payroll Summary"]

        # 2. Map your DataFrame (res_df) to the Template columns
        # We start at Row 10. 
        # Columns: B=2, C=3, D=4, I=9, J=10, K=11, L=12, M=13, N=14, O=15, Q=17
        start_row = 10
        
        for i, row_data in res_df.iterrows():
            current_excel_row = start_row + i
            
            # Fill the specific columns (Skipping E, F, G, H because of formulas)
            sheet.cell(row=current_excel_row, column=2).value = row_data.get('Staff No.', '')
            sheet.cell(row=current_excel_row, column=3).value = row_data.get('Employee Name', '')
            sheet.cell(row=current_excel_row, column=4).value = row_data.get('Basic Pay', 0)
            sheet.cell(row=current_excel_row, column=5).value = row_data.get('Other Allowances', 0)
            sheet.cell(row=current_excel_row, column=10).value = row_data.get('Pension Contribution', 0)
            sheet.cell(row=current_excel_row, column=11).value = row_data.get('NSSF', 0)
            sheet.cell(row=current_excel_row, column=12).value = row_data.get('SHIF', 0)
            sheet.cell(row=current_excel_row, column=13).value = row_data.get('AHL', 0)
            sheet.cell(row=current_excel_row, column=14).value = row_data.get('Taxable Salary', 0)
            sheet.cell(row=current_excel_row, column=15).value = row_data.get('Personal Relief', 0)
            sheet.cell(row=current_excel_row, column=16).value = row_data.get('PAYE', 0)
            sheet.cell(row=current_excel_row, column=17).value = row_data.get('Net Pay', 0)

        # 3. Optional: Hide extra rows if your template is longer than your data
        # If your template has 100 rows but you only used 15:
        max_template_rows = 100 
        for row_to_hide in range(start_row + len(res_df), max_template_rows):
            sheet.row_dimensions[row_to_hide].hidden = True

        # 4. Save to a Buffer for Streamlit Download
        output = io.BytesIO()
        wb.save(output)
        processed_data = output.getvalue()

        st.download_button(
            label="📊 Download Master Payroll",
            data=processed_data,
            file_name=f"{payroll_month}_Payroll_{payroll_year}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except FileNotFoundError:
        st.error("Template file not found! Make sure 'Payroll Summary Template.xlsx' is in your GitHub repo.")

    with d2:
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
            for _, row in res_df.iterrows():
                pdf_bytes = generate_payslip_pdf(row, payroll_month, payroll_year)
                zip_file.writestr(f"Payslip_{row['Staff No']}_{row['Employee']}.pdf", pdf_bytes)
        st.download_button("📂 Download All Payslips (ZIP)", zip_buffer.getvalue(), f"Payslips_{payroll_month}_{payroll_year}.zip")

    if st.button("🔄 Reset System"):
        st.session_state["employees"] = pd.DataFrame(columns=["Staff_No.", "Employee Name", "Basic Salary", "Other Allowances", "Pension Contribution"])
        st.session_state["results_df"] = None

        st.rerun()








